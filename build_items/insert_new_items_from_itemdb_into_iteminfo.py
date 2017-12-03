#!/usr/bin/env python.
'''
    File name: insert_new_items_from_itemdb_into_iteminfo.py
    Date created: December 2, 2017
    Python version: 3.6.1
    Purpose:
        Inserts item_db entries that do not exist in the iteminfo.lua
    Author: Phuc H Duong
    Website: phuchduong.io
    Linkedin: https://www.linkedin.com/in/phuchduong/
'''
import re  # regular expression
from os.path import isdir   # checks to see if a folder exists
import openpyxl  # excel plugin


def main():
    # for a list of valid codecs:
    #     essencero_restoration\Lua Codecs.xlsx
    # 437 is english codec
    encoding = "437"

    ###############
    # file system #
    ###############

    # your repo directory should look like this
    #
    # /repo root (essencero_restoration)
    #   \web_scrape_tamsinwhitfield
    #       /old_essence_item_db.txt
    #   \web_scrape_web_archive
    #       /item_db_web_archive.tsv
    if isdir("C:/repos"):
        repo_dir = "C:/repos"
    elif isdir("D:/repos"):
        repo_dir = "D:/repos"  # change this to your own

    #################################################################
    # 1. Parse in iteminfo.lua, formulate an item dictionary        #
    #################################################################
    iteminfo_dir = repo_dir + "/eRODev/eRO Client Data/system/itemInfosryx.lub"
    iteminfo_lua = parse_item_info_lua(
        file_dir=iteminfo_dir,
        item_dict={},
        encoding=encoding)
    #################################################################
    # 2. Parse in item_db, formulate an item dictionary             #
    #################################################################
    recon_dir = repo_dir + "/essencero_restoration/item_db_to_reconciliation/reconciliation.xlsx"
    recon_db = parse_reconciliation_spreadsheet(file_dir=recon_dir)
    #################################################################
    # 3. Insert entries that do not exist iteminfo.lua from item_db #
    #################################################################
    iteminfo_lua["iteminfo_db"] = insert_new_items_into_lua_db(
                                lua_db=iteminfo_lua["iteminfo_db"],
                                recon_db=recon_db)
    #################################################################
    # 4. Write out the item dictionary to a new iteminfo.lua        #
    #################################################################
    new_lua_dir = repo_dir + "/eRODev/eRO Client Data/system/new_itemInfosryx.lub"
    # write_lua_items_to_lua(file_dir=new_lua_dir, lua_parts=iteminfo_lua, encoding=encoding)


# Parses an iteminfo lua and returns a 3 element dictionary whoses keys are
# beg, mid, and end.
# beg is the begining of the file
# mid is the data structure in the middle of the file as a dictionary
# end is the remaining code in the file after the data structure
def parse_item_info_lua(file_dir, item_dict, encoding):
    print_opening_dir(file_dir=file_dir)
    ############
    # Patterns #
    ############
    # Begin item match
    new_item_line_pattern = "^\s{4,}\[\d{3,5}]\s{1,}?=\s{1,}?{$\n"
    is_new_item_line = re.compile(new_item_line_pattern)

    # 3~5 digit item code
    item_code_pattern = "\d{3,5}"
    extract_item_code = re.compile(item_code_pattern)

    # line that contains key value pairs
    key_value_line_pattern = '^\s{4,}\w{1,}\s{1,}?=\s{1,}?((".{0,}"|\d{1,})|{.{0,}}),?$\n'
    is_key_value_line = re.compile(key_value_line_pattern)

    # line that starts a multi line embedded object in the lua object
    multi_line_embed_key_pattern = '^\s{4,}\w{1,}\s{1,}?=\s{1,}?{$\n'
    is_multi_line_embed_key = re.compile(multi_line_embed_key_pattern)

    # line that has the values of a multi line embedded object
    multi_line_embed_value_pattern = '^\s{4,}(-- )?(\s{1,})?(\w{1,})?(\s{1,})?".{0,}",?$\n'
    is_multi_line_embed_value = re.compile(multi_line_embed_value_pattern)

    # marks the end of the data structure of the file by finding the main function
    end_of_data_pattern = "^function\s{1,}main\(.{0,}\)$\n|^main\s{0,}=\s{0,}function\(.{0,}\)$\n"
    is_end_of_data = re.compile(end_of_data_pattern)

    # return objects
    lua_beg = ""
    lua_end = "\n"

    # loop state objects
    current_item_id = -1
    stage_of_file = "beg"
    current_embed_key = ""

    with open(file=file_dir, mode="r", encoding=encoding) as lua:
        counter = 0
        for line in lua:
            if is_new_item_line.match(line):
                # if item code is found,
                # extract item code
                current_item_id = re.search(pattern=extract_item_code, string=line).group(0)
                if current_item_id not in item_dict:
                    item_dict[current_item_id] = {}
                    # make a new dictionary reference
                stage_of_file = "iteminfo_db"
            elif is_key_value_line.match(line):
                # if this is a key_value_pair line
                # add that key and value to the current item
                key_value_list = line.split("=")

                # key
                key = key_value_list[0].strip().replace("\n", "")

                # value
                value = key_value_list[1].strip().replace("\n", "")
                if value[-1:] == ",":  # removes trialing comma
                    # if there is a trailing comma, remove it
                    value = value[:-1]

                # adds key and value to the current item dict
                item_dict[current_item_id][key] = value
            elif is_multi_line_embed_key.match(line):
                # if it's the start of a multi line embed, create an embedded dictionary with a list
                # as it's value
                key = line.split("=")[0].strip()
                item_dict[current_item_id][key] = []
                current_embed_key = key
            elif is_multi_line_embed_value.match(line):
                # if it's the values of a multi line embed, append it to the descriptions list
                value = line.strip()
                if value[-1:] == ",":
                    value = value[:-1]
                item_dict[current_item_id][current_embed_key].append(value)
            elif stage_of_file == "beg":
                # if it's still part of the file before the item data structure,
                #   record everything to be rewritten later
                lua_beg += line
            elif is_end_of_data.match(line):
                # if it's past part of the item data structure in the file,
                #   record everything to be rewritten later
                stage_of_file = "end"
                lua_end += line
            elif stage_of_file == "end":
                # if it's past part of the item data structure in the file,
                #   record everything to be rewritten later
                lua_end += line
            else:
                # These are the lines skipped, print to be sure we didn't leave
                #     anything  behind.
                ignored_data = line.strip()
                if ignored_data != "":
                    if ignored_data[0] != "}":
                        print("Skipping line: " + str(counter) + ": " + line)
            counter += 1

    print("Parsed " + str(counter) + " lines from Lua.")
    item_lua_dict = {
        "beg": lua_beg,
        "iteminfo_db": item_dict,
        "end": lua_end
    }
    return item_lua_dict


# Tells the user in the console what file is currently being opened.
def print_opening_dir(file_dir):
    file_dir_split = file_dir.split("/")
    filename = file_dir_split[-1]
    file_path = file_dir_split[:-1]
    print("Opening: " + filename + " | From: " + "/".join(file_path))


# Tells the user how many lines were writte to a file
def print_writing_status(counter, file_dir):
    filename = file_dir.split("/")[-1]
    print("Found... " + str(counter) + " items in " + filename + "\n")


# Writes an iteminfo lua and from a 3 element dictionary parameter [1]<lua_parts> whoses keys are
# beg, mid, and end.
# beg is the begining of the file
# mid is the data structure in the middle of the file as a dictionary
# end is the remaining code in the file after the data structure
# [0]<file_dir> is the full path of the file to be written
# [2]<is_korean> can be True or False, true will specify encoding of ms949 for korean encoding
#     while false will yeild utf-8 encoding
def write_lua_items_to_lua(file_dir, lua_parts, encoding):
    print_opening_dir(file_dir=file_dir)

    # file parts
    lua_beg = lua_parts["beg"]
    lua_dict = lua_parts["iteminfo_db"]
    lua_end = lua_parts["end"]

    # writting specs
    spaces_per_tab = 4
    tab = " " * spaces_per_tab

    f = open(file=file_dir, mode="w+", encoding=encoding)
    f.write(lua_beg)  # first line

    lua_dict_keys = sorted([int(x) for x in lua_dict])
    for item_id in lua_dict_keys:
        item_id = str(item_id)
        f.write(tab + "[" + item_id + "] = {\n")
        for item_key in lua_dict[item_id]:
            if isinstance(lua_dict[item_id][item_key], list):

                multi_line_embed_str = tab * 2 + str(item_key) + " = {\n"

                for item in lua_dict[item_id][item_key]:
                    multi_line_embed_str += tab * 3 + item + ",\n"
                multi_line_embed_str += tab * 2 + "},\n"
                f.write(multi_line_embed_str)
            else:
                f.write(tab * 2 + str(item_key) + " = " + str(lua_dict[item_id][item_key]) + ",\n")
        f.write(tab + "},\n")
    f.write("}\n")
    f.write(lua_end)
    print("Writing " + str(len(lua_dict)) + " items to... " + file_dir)
    f.close()


# Takes in the reconciliation spreadsheet (xlsx) and parses it into an item dictionary
def parse_reconciliation_spreadsheet(file_dir):
    # to get values, use data_only=True, otherwise you'll get formulas
    excel = openpyxl.load_workbook(filename=file_dir, data_only=True)

    # to get sheetnames:
    # excel.get_sheet_names()
    target_sheet = "export"

    ex_sheet = excel.get_sheet_by_name(target_sheet)
    # Column reference
    c_type_name = 1
    # 2   Allowed Jobs
    # 3   Classes_String
    # 4   gender
    # 5   Loc
    c_description = 6
    c_sprite = 7
    # 8   old_id
    c_item_id = 9
    # 10  item key
    c_item_name = 10
    # 12  type
    # 13  price
    # 14  sell
    # 15  weight
    # 16  ATK[:MATK]
    # 17  DEF
    # 18  Range
    c_slot = 19
    # 20  Job
    # 21  Class
    # 22  Gender
    # 23  Loc
    # 24  wLV
    # 25  eLV[:maxLevel]
    # 26  Refineable
    c_view_id = 27
    # 28  { Script }
    # 29  { OnEquip_Script }
    # 30  { OnUnequip_Script }
    # 31  Concat
    item_dict = {}
    for i in range(1, ex_sheet.max_row + 1):
        type_name = ex_sheet.cell(row=i, column=c_type_name).value
        description = ex_sheet.cell(row=i, column=c_description).value
        sprite = ex_sheet.cell(row=i, column=c_sprite).value
        item_id = ex_sheet.cell(row=i, column=c_item_id).value
        item_name = ex_sheet.cell(row=i, column=c_item_name).value
        slot = ex_sheet.cell(row=i, column=c_slot).value
        view_id = ex_sheet.cell(row=i, column=c_view_id).value

        item_dict[item_id] = {
            "type_name": type_name,
            "description": description,
            "sprite": sprite,
            "item_name": item_name,
            "slot": slot,
            "view_id": view_id,
        }
    return item_dict


# Adds to the lua db, entries that do not exist in the lua db
# from the recon db.
def insert_new_items_into_lua_db(lua_db,recon_db):
    for item_id in recon_db:
        item_id = str(item_id)
        if item_id in lua_db:
            # Item already exists in the iteminfo
            # In the future, use this area to override bad descriptions
            print("Found... " + item_id + "Skipping...")
        elif item_id not in lua_db:
            # Insert new items into the item description

            lua_db[item_id]["unidentifiedDisplayName"] = get_unidentifiedDisplayName(item_entry=lua_db[])
            lua_db[item_id]["unidentifiedResourceName"] = get_unidentifiedResourceName(item_entry=lua_db[])
            lua_db[item_id]["unidentifiedDescriptionName"] = get_unidentifiedDescriptionName(item_entry=lua_db[])
            lua_db[item_id]["identifiedDisplayName"] = get_identifiedDisplayName(item_entry=lua_db[])
            lua_db[item_id]["identifiedResourceName"] = get_identifiedResourceName(item_entry=lua_db[])
            lua_db[item_id]["identifiedDescriptionName"] = get_identifiedDescriptionName(item_entry=lua_db[])
            lua_db[item_id]["slotCount"] = get_slotCount(item_entry=lua_db[])
            lua_db[item_id]["ClassNum"] = get_ClassNum(item_entry=lua_db[])
        else:
            print("Something went wrong:" + item_id)
    return lua_db



# Derives unidentifiedDisplayName from the item entry
def get_unidentifiedDisplayName(item_entry)
    return unidentifiedDisplayName


# Derives unidentifiedResourceName from the item entry
def get_unidentifiedResourceName(item_entry)
    return unidentifiedResourceName


# Derives unidentifiedDescriptionName from the item entry
def get_unidentifiedDescriptionName(item_entry)
    return unidentifiedDescriptionName


# Derives identifiedDisplayName from the item entry
def get_identifiedDisplayName(item_entry)
    return identifiedDisplayName


# Derives identifiedResourceName from the item entry
def get_identifiedResourceName(item_entry)
    return identifiedResourceName


# Derives identifiedDescriptionName from the item entry
def get_identifiedDescriptionName(item_entry)
    return identifiedDescriptionName


# Derives slotCount from the item entry
def get_slotCount(item_entry)
    return slotCount


# Derives ClassNum from the item entry
def get_ClassNum(item_entry)
    return ClassNum



# Takes in a dictionary and creates a list of headers for a csv file
# based upon all keys inside of the dictionary.
# Params:
#   dict: takes in a dictioanry to populate the headers with its keys
#   prefix: what the items headers should be prefixed with, helps mitigate colisions when
#       merging two dictionaries together
#   pk: primary key name of the parent key of the file to be incorporated into the list of header names
# Return: a list of header names
def scan_headers(dictionary, name_of_pk):
    headers = [name_of_pk]
    for item_id in dictionary:  # loop over all items
        for attribute in dictionary[item_id]:  # loop over all attributes of items
            if attribute not in headers:
                headers.append(attribute)
    return headers


main()
